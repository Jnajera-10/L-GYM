import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO

class ExportService:
    @staticmethod
    def _header_style(ws, headers):
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1a1a2e")
            cell.alignment = Alignment(horizontal="center")

    @staticmethod
    def export_clients_excel(clients):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Clientes'
        ExportService._header_style(ws, ['ID', 'Nombre Completo', 'Tipo Doc', 'Número Doc', 'Email', 'Celular', 'Estado'])
        for c in clients:
            ws.append([c.id, c.full_name, c.document_type, c.document_number,
                       c.email or '', c.phone or '', 'Activo' if c.is_active else 'Inactivo'])
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 18
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    @staticmethod
    def export_payments_excel(payments):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Pagos'
        ExportService._header_style(ws, ['ID', 'Cliente', 'Membresía', 'Monto', 'Fecha Pago', 'Vencimiento', 'Método'])
        for p in payments:
            ws.append([p.id,
                       p.client.full_name if p.client else 'N/A',
                       p.membership.name if p.membership else 'N/A',
                       p.amount,
                       str(p.payment_date),
                       str(p.end_date),
                       p.payment_method])
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 18
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    @staticmethod
    def export_sales_excel(sales):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Ventas'
        ExportService._header_style(ws, ['ID', 'Cliente', 'Total', 'Método Pago', 'Fecha', 'Nº Productos'])
        for s in sales:
            ws.append([s.id,
                       s.client.full_name if s.client else 'Cliente general',
                       s.total,
                       s.payment_method,
                       s.sale_date.strftime('%d/%m/%Y %H:%M'),
                       len(s.items)])
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 18
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf
